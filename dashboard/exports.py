"""
Export utilities: CSV, Excel, PDF.
"""
import csv
import io
from django.http import HttpResponse
from core.utils import format_duration


class CSVExporter:
    """Export leaderboard data as CSV."""

    @staticmethod
    def export(quiz, entries):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{quiz.title}_results.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Rank', 'Name', 'College', 'Score', 'Accuracy (%)',
            'Time Taken', 'Questions Attempted', 'Correct', 'Wrong'
        ])

        for entry in entries:
            writer.writerow([
                entry['rank'],
                entry['name'],
                entry['college'],
                entry['score'],
                f"{entry['accuracy']}%",
                format_duration(entry['time_taken']),
                entry['questions_attempted'],
                entry['correct_answers'],
                entry['wrong_answers'],
            ])

        return response


class ExcelExporter:
    """Export leaderboard data as Excel (.xlsx)."""

    @staticmethod
    def export(quiz, entries):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Results"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")

        headers = [
            'Rank', 'Name', 'College', 'Score', 'Accuracy (%)',
            'Time Taken', 'Questions Attempted', 'Correct', 'Wrong'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row, entry in enumerate(entries, 2):
            ws.cell(row=row, column=1, value=entry['rank'])
            ws.cell(row=row, column=2, value=entry['name'])
            ws.cell(row=row, column=3, value=entry['college'])
            ws.cell(row=row, column=4, value=entry['score'])
            ws.cell(row=row, column=5, value=f"{entry['accuracy']}%")
            ws.cell(row=row, column=6, value=format_duration(entry['time_taken']))
            ws.cell(row=row, column=7, value=entry['questions_attempted'])
            ws.cell(row=row, column=8, value=entry['correct_answers'])
            ws.cell(row=row, column=9, value=entry['wrong_answers'])

        # Auto-width columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = max_length

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{quiz.title}_results.xlsx"'
        wb.save(response)
        return response


class PDFExporter:
    """Export leaderboard data as PDF."""

    @staticmethod
    def export(quiz, entries):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph(f"<b>{quiz.title} — Results</b>", styles['Title']))
        elements.append(Spacer(1, 20))

        # Table data
        data = [[
            'Rank', 'Name', 'College', 'Score', 'Accuracy',
            'Time', 'Attempted', 'Correct', 'Wrong'
        ]]

        for entry in entries:
            data.append([
                str(entry['rank']),
                entry['name'],
                entry['college'],
                str(entry['score']),
                f"{entry['accuracy']}%",
                format_duration(entry['time_taken']),
                str(entry['questions_attempted']),
                str(entry['correct_answers']),
                str(entry['wrong_answers']),
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C63FF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))

        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{quiz.title}_results.pdf"'
        return response
